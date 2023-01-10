import datetime
from decimal import Decimal
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from openpyxl import Workbook

from centroscustos.models import CentroCustosModel
from lancamentos.models import LancamentosModel
from previsoes.models import ClassificacaoDespesas
from django.http import JsonResponse
from relatorios.formularios import formularoRelatorioLancamentos
from django.db.models import Q, Sum, Prefetch
from weasyprint import HTML
from weasyprint.fonts import FontConfiguration
from django.template.loader import render_to_string

def resultadoBusca(request):
    if request.method=='POST':
        if 'sessaoLancamentos' in request.session:
            del request.session['sessaoLancamentos']
        form = formularoRelatorioLancamentos(request.POST)
        if form.is_valid():
            descricao=form.cleaned_data['descricao']
            centrosdeCustos=form.cleaned_data['centroCusto']
            datainicial=form.cleaned_data['dataInicial']
            dataFinal=form.cleaned_data['dataFinal']
            classificacao=form.cleaned_data['classificacao']
            qs = LancamentosModel.objects.filter(status=True).filter(data_pagamento__range=[datainicial, dataFinal]).order_by('data_pagamento')
            if descricao:
                descricao = request.GET.get('descricao')
                qs = qs.filter(descricao__icontains=descricao).order_by('operacao', 'centro_custos__titulo',
                                                                        'classificacao_despesa__titulo', 'data_pagamento')
            listaCentrosCustos = []
            for item in centrosdeCustos:
                objeto = CentroCustosModel.objects.get(pk=item.pk)
                listaCentrosCustos.append(objeto.titulo)
            if len(listaCentrosCustos)>0:
                qs = qs.filter(Q(centro_custos__titulo__in=listaCentrosCustos)).order_by('operacao',
                                                                                         'centro_custos__titulo',
                                                                                         'classificacao_despesa__titulo',
                                                                                         'data_pagamento')
            listaClassificacoes =[]
            for item in classificacao:
                objeto=ClassificacaoDespesas.objects.get(pk=item.pk)
                listaClassificacoes.append(objeto.titulo)
            if len(listaClassificacoes) > 0:
                qs = qs.filter(Q(classificacao_despesa__titulo__in=listaClassificacoes)).order_by('operacao',
                                                                                                      'centro_custos__titulo',
                                                                                                      'classificacao_despesa__titulo',
                                                                                                      'data_pagamento')
            somatorio=qs.aggregate(total=Sum('valor_final'))
            acumulado=somatorio['total']
            if somatorio['total'] is None:
                acumulado = Decimal(0)
            valores = qs.values('descricao', 'classificacao_despesa__titulo', 'centro_custos__titulo', 'operacao',
                                'previsao__titulo', 'valor_final', 'data_pagamento')
            contexto={
                'object_list': valores,
                'itensencontrados':qs.count(),
                'centrosdeCustos':centrosdeCustos,
                'datainicial':datainicial,
                'datafinal':dataFinal,
                'classificacaodespesas':classificacao,
                'valortotal':acumulado
            }
            sessaoLancamentos=[item.id for item in qs]
            sessaoListaCentrosCustos=[item for item in centrosdeCustos]
            #sessaoClassificacoes=[item for item in valores['classificacao_despesa__titulo']]

            request.session['sessaoLancamentos']=sessaoLancamentos
            """
            request.session['sessaoListaCentrosCustos']=sessaoListaCentrosCustos
            request.session['sessaoClassificacao']=sessaoClassificacoes
            request.session['sessaoDataInicial']=datainicial
            request.session['sessaoDataFinal']=dataFinal
            """
            return render(request, 'relatorios/01/lista.html', contexto)
        else:
            for item in form.errors:
                messages.add_message(request, messages.ERROR, item[1])
            return redirect(request.META['HTTP_REFERER'])
def busca(request):
    contexto={
        'formulario':formularoRelatorioLancamentos
    }
    return render(request, 'relatorios/01/busca.html', contexto)
def novoRelatorio(request):
    del request.session['sessaoLancamentos']
    messages.add_message(request, messages.SUCCESS, 'Você já pode gerar um novo relatório')
    return redirect('relatorios_busca')

def gerarPDFRelatorio(request):
    if 'sessaoLancamentos' not in request.session:
        messages.add_message(request, messages.WARNING, 'Sem dados para a emissão do relatório')
        messages.add_message(request, messages.WARNING,
                             'Você precisa definir os parâmetros antes de solicitar a exportação de um relatório')
        return redirect('relatorios_busca')
    else:
        sessaoLancamentos = request.session['sessaoLancamentos']

        qs=LancamentosModel.objects.filter(id__in=sessaoLancamentos)

        categorias=ClassificacaoDespesas.objects.filter(lancamentos_classificacaodespesa__in=qs).\
            annotate(total=Sum('lancamentos_classificacaodespesa__valor_final'))\
            .prefetch_related(Prefetch('lancamentos_classificacaodespesa',
                                       LancamentosModel.objects.filter(id__in=qs).annotate(total=Sum('valor_final')),
                                       to_attr='lacamentos_valore_final'))
        teste=categorias.count()
        lista = qs.values('descricao', 'classificacao_despesa__titulo', 'centro_custos__titulo', 'operacao',
                            'previsao__titulo', 'valor_final', 'data_pagamento').order_by('classificacao_despesa__titulo', 'data_pagamento')
        totalEntradas=qs.filter(valor_final__gt=0).aggregate(total=Sum('valor_final'))
        while totalEntradas['total'] == '' or totalEntradas['total'] == None:
            totalEntradas['total']=Decimal(0.00)

        totalSaidas=qs.filter(valor_final__lt=0).aggregate(total=Sum('valor_final'))
        while totalSaidas['total'] == '' or totalSaidas['total'] == None:
            totalSaidas['total']=Decimal(0.00)
        totalFinal=qs.aggregate(total=Sum('valor_final'))
        while totalFinal['total'] == '' or totalFinal['total'] == None:
            totalFinal['total']=Decimal(0.00)

        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = "inline; filename=relatorio-lancamentos-{date}.pdf".format(
            date=datetime.datetime.now(),
        )
        html = render_to_string("relatorios/lancamentos.html", {
            'object_list': lista,
            'teste': categorias,
            'totalEntrada': totalEntradas['total'],
            'totalSaida': totalSaidas['total'],
            'totalFinal': totalFinal['total'],
        })
        font_config=FontConfiguration()
        HTML(string=html).write_pdf(response, font_config=font_config)
        return response

def gerarXSLRelatorio(request):
    if 'sessaoLancamentos' not in request.session:
        messages.add_message(request, messages.WARNING, 'Sem dados para a emissão do relatório')
        messages.add_message(request, messages.WARNING,
                             'Você precisa definir os parâmetros antes de solicitar a exportação de um relatório')
        return redirect('relatorios_busca')
    else:
        sessaoLancamentos = request.session['sessaoLancamentos']

        qs = LancamentosModel.objects.filter(id__in=sessaoLancamentos)

        categorias = ClassificacaoDespesas.objects.filter(lancamentos_classificacaodespesa__in=qs). \
            annotate(total=Sum('lancamentos_classificacaodespesa__valor_final')) \
            .prefetch_related(Prefetch('lancamentos_classificacaodespesa',
                                       LancamentosModel.objects.filter(id__in=qs).annotate(total=Sum('valor_final')),
                                       to_attr='lacamentos_valore_final'))
        teste = categorias.count()
        lista = qs.values('descricao', 'classificacao_despesa__titulo', 'centro_custos__titulo', 'operacao',
                          'previsao__titulo', 'valor_final', 'data_pagamento').order_by('classificacao_despesa__titulo',
                                                                                        'data_pagamento')
        totalEntradas = qs.filter(valor_final__gt=0).aggregate(total=Sum('valor_final'))
        while totalEntradas['total'] == '' or totalEntradas['total'] == None:
            totalEntradas['total'] = Decimal(0.00)

        totalSaidas = qs.filter(valor_final__lt=0).aggregate(total=Sum('valor_final'))
        while totalSaidas['total'] == '' or totalSaidas['total'] == None:
            totalSaidas['total'] = Decimal(0.00)
        totalFinal = qs.aggregate(total=Sum('valor_final'))
        while totalFinal['total'] == '' or totalFinal['total'] == None:
            totalFinal['total'] = Decimal(0.00)

        wb=Workbook()
        planilha=wb.active
        planilha.title=str('Relatorio-de-lancamentos')
        planilha['A1'] = "#"
        planilha['B1'] = "TÍTULO"
        planilha['C1'] = "CENTRO DE CUSTOS"
        planilha['D1'] = "CLASSIFICAÇÃO"
        planilha['E1'] = "DATA"
        planilha['F1'] = "VALOR"
        index = 2
        numero_linha = 1
        for linha in lista:
            numero='#'
            titulo=linha['descricao']
            cc=linha['centro_custos__titulo']
            classificacao=linha['classificacao_despesa__titulo']
            data=linha['data_pagamento']
            valor=linha['valor_final']
            planilha.cell(column=1, row=index, value=numero_linha)
            planilha.cell(column=2, row=index, value=titulo)
            planilha.cell(column=3, row=index, value=cc)
            planilha.cell(column=4, row=index, value=classificacao)
            planilha.cell(column=5, row=index, value=data)
            planilha.cell(column=6, row=index, value=valor)
            index += 1
            numero_linha += 1
        planilha.cell(column=4, row=index, value=str('Total em entradas {}'.format(totalEntradas['total'])))
        planilha.cell(column=5, row=index, value=str('Total em saídas {}'.format(totalSaidas['total'])))
        planilha.cell(column=6, row=index, value=str('Total {}'.format(totalFinal['total'])))
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=relatorio-lancamentos-{}.xlsx'.format(datetime.datetime.now())
        wb.save(response)
        return response