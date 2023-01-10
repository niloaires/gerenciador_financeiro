from django.shortcuts import render, get_object_or_404
import datetime
# Create your views here.
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.text import slugify
from weasyprint import HTML
from weasyprint.fonts import FontConfiguration
from lancamentos.models import LancamentosModel, FornecedoresLancamentosModel
from fornecedores.models import FornecedoresModel
from contas.models import ContasModel
from centroscustos.models import CentroCustosModel
from previsoes.models import PrevisoesModel, ParcelasModel, PrevisoesFornecedoresModel, ClassificacaoDespesas
from obras.models import ObrasModel, LancamentosObrasModel
from orcamentos.models import *
from clientes.models import *
from django.db.models import Sum, Q, Prefetch, Count, Case, When
import locale
from django.contrib.auth.decorators import login_required
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
from core.decorador import precisa_ser_gestor
from relatorios.formularios import FormPesquisaGeral
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
@login_required(login_url='login')
@precisa_ser_gestor
def lancamentos_pdf(request):
    object_list=LancamentosModel.objects.filter(status=True).order_by('-data_registro')
    acumulado= object_list.aggregate(total=Sum('valor_final'))
    response=HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = "inline; filename=relatorio-lancamentos-{date}.pdf".format(
        date=datetime.datetime.now(),
    )
    html = render_to_string("relatorios/lancamentos.html", {
        'object_list': object_list,
        'data': datetime.datetime.now(),
        'acumulado': acumulado['total']
    })
    font_config = FontConfiguration()
    HTML(string=html).write_pdf(response, font_config=font_config)
    return response
@login_required(login_url='login')
@precisa_ser_gestor
def lancamento_especifico(request, uuid):
    object=LancamentosModel.objects.get(uuid=uuid)
    contexto = {
        'object': object,
        'data': datetime.datetime.now(),
    }
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = "inline; filename=relatorio-lancamento-{object}-{date}.pdf".format(
        object=object.descricao,
        date=datetime.datetime.now())
    html = render_to_string("relatorios/lancamento.html", contexto)
    font_config = FontConfiguration()
    HTML(string=html).write_pdf(response, font_config=font_config)
    return response
@login_required(login_url='login')
@precisa_ser_gestor
def fornecedores_pdf(request):
    object_list=FornecedoresModel.objects.filter(status=True)\
        .prefetch_related('fornecedores_lancamento')\
        .values('nome', 'tipo_fornecedor', 'tipo_conta',
                'agencia', 'conta', 'data_registro',
                'fornecedores_lancamento__data_registro',
                'fornecedores_lancamento__lancamento__valor_final',
                'banco', 'cpf_cnpj', 'pix')\
        .annotate(total=Sum('fornecedores_lancamento__lancamento__valor_final'))\
        .order_by('nome')
    acumulado = object_list.aggregate(total=Sum('fornecedores_lancamento__lancamento__valor_final'))
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = "inline; filename=relatorio-fornecedores-{date}.pdf".format(
        date=datetime.datetime.now(),
    )
    html = render_to_string("relatorios/fornecedores.html", {
        'object_list': object_list,
        'data': datetime.datetime.now(),
        'acumulado': acumulado['total']
    })
    font_config = FontConfiguration()
    HTML(string=html).write_pdf(response, font_config=font_config)
    return response
@login_required(login_url='login')
@precisa_ser_gestor
def contas_a_apagar_pdf(request):
    object_list=PrevisoesFornecedoresModel.\
        objects.select_related('fornecedor', 'previsao').\
        prefetch_related('previsao__parcelas_previsao').\
        values('fornecedor__nome','previsao__codigo', 'previsao__referencia', 'previsao__centro_custos__titulo',
               'previsao__parcelas_previsao__titulo',
               'previsao__parcelas_previsao__efetivada',
               'previsao__parcelas_previsao__data_vencimento' ,
               'previsao__parcelas_previsao__valor')

    acumulado = object_list.aggregate(total=Sum('previsao__parcelas_previsao__valor'))
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = "inline; filename=relatorio-contas-a-pagar-{date}.pdf".format(
        date=datetime.datetime.now())
    html = render_to_string("relatorios/contas_a_pagar.html", {
        'object_list': object_list,
        'data': datetime.datetime.now(),
        'acumulado': acumulado['total']
    })
    font_config = FontConfiguration()
    HTML(string=html).write_pdf(response, font_config=font_config)
    return response
@login_required(login_url='login')
@precisa_ser_gestor
def relatorio_centro_custos(request):
    qs=CentroCustosModel.objects.filter(status=True)
    lancamentos=qs.prefetch_related('lancamentos_centrocusto').\
        values('titulo','lancamentos_centrocusto__descricao',
               'lancamentos_centrocusto__previsao__codigo',
               'lancamentos_centrocusto__data_pagamento',
               'lancamentos_centrocusto__previsao__efetivada',
               'lancamentos_centrocusto__valor_final', 'usuario__first_name', 'usuario__last_name').order_by('titulo', 'lancamentos_centrocusto__data_pagamento')
    previsoes=PrevisoesModel.objects.select_related('centro_custos').\
        values('titulo', 'centro_custos__titulo', 'codigo',
               'data_previsao', 'usuario__first_name',
               'usuario__last_name', 'valor').filter(efetivada=False)
    contexto={
        'lancamentos':lancamentos,
        'previsoes':previsoes,
        'data': datetime.datetime.now(),
    }
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = "inline; filename=relatorio-contas-a-pagar-{date}.pdf".format(
        date=datetime.datetime.now())
    html = render_to_string("relatorios/centros_custo.html", contexto)
    font_config = FontConfiguration()
    HTML(string=html).write_pdf(response, font_config=font_config)
    return response
@login_required(login_url='login')
@precisa_ser_gestor
def relatorio_centro_custos_especifico(request, uuid):
    if request.method == 'POST':
        data_inicio = request.POST.get('data_inicio')
        data_final = request.POST.get('data_fim')
        formato = request.POST.get('formato_relatorio')
    object = CentroCustosModel.objects.get(uuid=uuid)
    if request.method == 'POST':
        object_list = LancamentosModel.objects.filter(centro_custos__uuid=uuid,
                                                      data_pagamento__range=[data_inicio, data_final],
                                                      status=True)

        acumulado = LancamentosModel.objects.filter(centro_custos__uuid=uuid,
                                                    data_pagamento__range=[data_inicio, data_final],
                                                    status=True).aggregate(
            total=Sum('valor_final'))
    else:
        object_list = ClassificacaoDespesas.objects.filter(lancamentos_classificacaodespesa__centro_custos=object,
                                                           lancamentos_classificacaodespesa__data_pagamento__range=[
                                                               data_inicio, data_final]). \
            annotate(total=Sum('lancamentos_classificacaodespesa__valor_final')). \
            prefetch_related(Prefetch('lancamentos_classificacaodespesa',
                                      LancamentosModel.objects.annotate(total=Sum('valor_final')),
                                      to_attr='itens_subtotais')
                             )
        acumulado = LancamentosModel.objects.filter(status=True, centro_custos=object).aggregate(
            total=Sum('valor_final'))
    contexto = {
        'object': object,
        'object_list': object_list,
        'resumo': ClassificacaoDespesas.objects.prefetch_related('previsoes_classificacaodespesa').values('titulo',
                                                                                                          'previsoes_classificacaodespesa__titulo',
                                                                                                          'previsoes_classificacaodespesa__valor'),
        'data': datetime.datetime.now(),
        'acumulado': acumulado['total']
    }

    if formato == 'excel':
        wb = Workbook()
        planilha = wb.active
        planilha.title = str(object.titulo)
        planilha['A1'] = "#"
        planilha['B1'] = "Descrição"
        planilha['C1'] = "Classificação"
        planilha['D1'] = "Data do pagamento"
        planilha['E1'] = "Responsável"
        planilha['F1'] = "Entrada/Saída"
        planilha['G1'] = "Valor"
        index = 2
        numero_linha = 1
        for linha in object_list:
            numero = "#"
            descricao = linha.descricao
            classificacao = linha.previsao.classificacao_despesa.titulo
            datapagamento = str(linha.previsao.data_pagamento)
            responsavel = linha.previsao.usuario.first_name
            valor = linha.previsao.valor
            entradasaida = linha.previsao.get_modalidade_display()
            planilha.cell(column=1, row=index, value=numero_linha)
            planilha.cell(column=2, row=index, value=descricao)
            planilha.cell(column=3, row=index, value=classificacao)
            planilha.cell(column=4, row=index, value=datapagamento)
            planilha.cell(column=5, row=index, value=responsavel)
            planilha.cell(column=6, row=index, value=entradasaida)
            planilha.cell(column=7, row=index, value=valor)
            index += 1
            numero_linha += 1
            # Save the workbook

        planilha.cell(column=7, row=index, value=acumulado['total'])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=export'
        wb.save(response)
        return response
    else:
        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = "inline; filename=relatorio-centro-custo {object}-{date}.pdf".format(object=object.titulo,
            date=datetime.datetime.now())
        html = render_to_string("relatorios/centros_custo_especifico-bkp.html", contexto)
        font_config = FontConfiguration()
        HTML(string=html).write_pdf(response, font_config=font_config)
        return response


@login_required(login_url='login')
@precisa_ser_gestor
def relatorio_conta_bancaria_especifica(request, uuid, inicial, final):
    object=ContasModel.objects.get(uuid=uuid)
    qs=LancamentosModel.objects.filter(status=True)
    if inicial and final:
        qs.filter(data_pagamento__range=[inicial, final])
    object_list=qs
    acumulado = object_list.aggregate(total=Sum('valor'))
    contexto = {
        'object': object,
        'object_list': object_list,
        'data': datetime.datetime.now(),
        'acumulado': locale.currency(acumulado['total'], symbol=True, grouping=True)
    }
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = "inline; filename=relatorio-centro-custo {object}-{date}.pdf".format(
        object=object.titulo,
        date=datetime.datetime.now())
    html = render_to_string("relatorios/centros_custo_especifico.html", contexto)
    font_config = FontConfiguration()
    HTML(string=html).write_pdf(response, font_config=font_config)
    return response
@login_required(login_url='login')
@precisa_ser_gestor
def relatorio_obra_especifica(request, uuid):
    object=ObrasModel.objects.get(uuid=uuid)
    qs=LancamentosObrasModel.objects.filter(obra__uuid=uuid)

    object_list=qs
    acumulado = object_list.aggregate(total=Sum('lancamento__valor_final'))
    contexto = {
        'object': object,
        'object_list': object_list,
        'data': datetime.datetime.now(),
        'acumulado': locale.currency(acumulado['total'], symbol=True, grouping=True)
    }
    wb = Workbook()
    planilha = wb.active
    planilha.title = str(object.nome)
    planilha['A1'] = "#"
    planilha['B1'] = "Descrição"
    planilha['C1'] = "Centro de custos"
    planilha['D1'] = "Data do pagamento"
    planilha['E1'] = "Responsável"
    planilha['F1'] = "Entrada/Saída"
    planilha['G1'] = "Valor"
    index = 2
    numero_linha = 1
    for linha in object_list:
        numero = "#"
        cc = linha.lancamento.centro_custos.titulo
        datapagamento = str(linha.lancamento.data_pagamento)
        responsavel = linha.lancamento.usuario.first_name
        valor = linha.lancamento.valor_final
        entradasaida = linha.previsao.get_modalidade_display()
        planilha.cell(column=1, row=index, value=numero_linha)
        planilha.cell(column=2, row=index, value=linha)
        planilha.cell(column=3, row=index, value=cc)
        planilha.cell(column=4, row=index, value=datapagamento)
        planilha.cell(column=5, row=index, value=responsavel)
        planilha.cell(column=6, row=index, value=entradasaida)
        planilha.cell(column=7, row=index, value=valor)
        index += 1
        numero_linha += 1
        # Save the workbook

    planilha.cell(column=7, row=index, value=acumulado['total'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=export'
    wb.save(response)
    return response
    #response = HttpResponse(content_type="application/pdf")
    #response['Content-Disposition'] = "inline; filename=relatorio-obra-{object}-{date}.pdf".format(
    #    object=object.nome,
    #    date=datetime.datetime.now())
    #html = render_to_string("relatorios/obra_especifica.html", contexto)
    #font_config = FontConfiguration()
    #HTML(string=html).write_pdf(response, font_config=font_config)
    #return response
@login_required(login_url='login')
@precisa_ser_gestor
def relatorio_cliente_especifico(request, uuid):
    object=ClientesModel.objects.get(uuid=uuid)
    qs=CLientesLancamentosModel.objects.filter(cliente__uuid=uuid, lancamento__status=True)
    qs.values('cliente__cliente_contrato_lancamento__contrato')
    object_list=qs
    acumulado = object_list.aggregate(total=Sum('lancamento__valor_final'))
    contexto = {
        'object': object,
        'object_list': object_list,
        'data': datetime.datetime.now(),
        'acumulado': locale.currency(acumulado['total'], symbol=True, grouping=True)
    }
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = "inline; filename=relatorio-cliente-{object}-{date}.pdf".format(
        object=object.nome,
        date=datetime.datetime.now())
    html = render_to_string("relatorios/cliente_especifico.html", contexto)
    font_config = FontConfiguration()
    HTML(string=html).write_pdf(response, font_config=font_config)
    return response
@login_required(login_url='login')
@precisa_ser_gestor
def previsa_especifica(request, uuid):
    object=PrevisoesModel.objects.get(uuid=uuid)
    contexto = {
        'object': object,
        'data': datetime.datetime.now(),
    }
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = "inline; filename=relatorio-previsao-{object}-{date}.pdf".format(
        object=object.titulo,
        date=datetime.datetime.now())
    html = render_to_string("relatorios/previsao.html", contexto)
    font_config = FontConfiguration()
    HTML(string=html).write_pdf(response, font_config=font_config)
    return response
@login_required(login_url='login')
def os_especifica(request, uuid):
    object=OrcamentosModel.objects.get(uuid=uuid)
    contexto = {
        'object': object,
        'data': datetime.datetime.now(),
    }
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = "inline; filename=relatorio-ordem_de_servico-{object}-{date}.pdf".format(
        object=object,
        date=datetime.datetime.now())
    html = render_to_string("relatorios/os_especifico.html", contexto)
    font_config = FontConfiguration()
    HTML(string=html).write_pdf(response, font_config=font_config)
    return response
ClassificacaoDespesas.objects.prefetch_related('previsoes_classificacaodespesa__lancamentos_previsao')
def Teste(request):
    dados= ClassificacaoDespesas.objects.filter(lancamentos_classificacaodespesa__centro_custos=object).\
        annotate(total=Sum('lancamentos_classificacaodespesa__valor_final')).\
        prefetch_related(Prefetch('lancamentos_classificacaodespesa',
                                  LancamentosModel.objects.annotate(total=Sum('valor_final')),
                                  to_attr='itens_subtotais')
                         )
    return render(request, 'teste.html', {'dados':dados})
@login_required(login_url='login')
@precisa_ser_gestor
def relatorio__previsao__centro_custos(request, uuid):
    data_inicio=request.POST.get('data_inicio')
    data_final=request.POST.get('data_fim')

    object=CentroCustosModel.objects.get(uuid=uuid)
    object_list = ClassificacaoDespesas.objects.filter(previsoes_classificacaodespesa__centro_custos=object,
                                                       previsoes_classificacaodespesa__efetivada=False,
                                                       previsoes_classificacaodespesa__data_previsao__range=[data_inicio, data_final]). \
        annotate(total=Sum('previsoes_classificacaodespesa__valor')). \
        prefetch_related(Prefetch('previsoes_classificacaodespesa',
                                  PrevisoesModel.objects.annotate(total=Sum('valor')),
                                  to_attr='itens_subtotais')
                         )

    acumulado=PrevisoesModel.objects.filter(status=True, efetivada=False, data_previsao__range=[data_inicio, data_final]).aggregate(total=Sum('valor'))
    contexto={
        'object':object,
        'object_list':object_list,
        'resumo':ClassificacaoDespesas.objects.prefetch_related('previsoes_classificacaodespesa').values('titulo', 'previsoes_classificacaodespesa__titulo', 'previsoes_classificacaodespesa__valor'),
        'data': datetime.datetime.now(),
        'acumulado':acumulado['total']
    }
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = "inline; filename=relatorio-centro-custo {object}-{date}.pdf".format(object=object.titulo,
        date=datetime.datetime.now())
    html = render_to_string("relatorios/relatorio__previsao__centro_custos.html", contexto)
    font_config = FontConfiguration()
    HTML(string=html).write_pdf(response, font_config=font_config)
    return response
@login_required(login_url='login')
@precisa_ser_gestor
def view_relatorio(request):
    contexto={
        'formulario':FormPesquisaGeral
    }
    return render(request, 'relatorios/pesquisa_geral.html', contexto)


